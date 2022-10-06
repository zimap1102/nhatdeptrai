#!/usr/bin/env python
import sys
from openpyxl import load_workbook

wb = load_workbook(filename = sys.argv[1])
wb.create_sheet(title='sheet')
wb.save('%s.out' % sys.argv[1])
